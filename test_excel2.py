from openpyxl import Workbook

# Excelドキュメントを作成
wb = Workbook()

# 新しいワークシートを追加
ws = wb.create_sheet(title="New Sheet")

# 10行×10列のセルにインクリメントした数字を設定
count = 1
for row in range(1, 11):
    for col in range(1, 11):
        ws.cell(row=row, column=col, value=count)
        count += 1

# Excelドキュメントを保存
wb.save("sample.xlsx")